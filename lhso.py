import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === Step 1: Baca dan proses cepat pakai pandas ===
df = pd.read_excel('LHSO.xlsx', header=None)  # Treating the first row as data, not headers

# Pilih kolom yang diinginkan berdasarkan indeks kolom
columns_to_keep = [0, 4,22,37,41,49,54,60,65,73,77]  # These are the indices of A, F, X, AM, AQ, AY, BD, BJ, BO, BW, CA

df = df.iloc[:, columns_to_keep]

# Hapus baris yang kolom pertama-nya kosong
df = df[df.iloc[:, 0].notna()].copy()

# Tambahkan nomor urut di kolom pertama
df.insert(0, 'No', range(1, len(df) + 1))

# Simpan hasil sementara
df.to_excel('file_diedit.xlsx', index=False, header=False)

# === Step 2: Buka hasil dan atur lebar kolom (jika mau) ===
wb = load_workbook('file_diedit.xlsx')
ws = wb.active

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.save('file_diedit.xlsx')
